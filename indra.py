from openpyxl import load_workbook


def preencher():
    a = ler_clientes()
    dados = a[0]
    celulas = a[1]
    transcrever(dados, celulas)

def ler_clientes():
    # leitura do arquivo excel
    wb = load_workbook('Cliente.xlsx')

    #criando a variável ws (variável python que armazena os dados da planilha)
    ws = []
    for sheet in wb:
        ws = sheet

    #criando as variáveis python que irão receber os valores do excel
    dados = []
    celulas = []

    #extraindo os valores da planilha para as variáveis python
    for x in range(1,len(tuple(ws.rows))+1):
        dados.append(ws['B'+ str(x)].value)
        celulas.append(ws['C'+str(x)].value)

    return dados, celulas


def transcrever(dados, celulas):
    #leitura do arquivo
    wb = load_workbook('NT.020.EQTL.Normas e Padrões - 03 - Anexo I _ Formulário de Solicitação de Acesso para Microgeração Distribuída até 10 kW.xlsx')

    #atribuindo variáveis a cada sheet para facilitar o seu uso
    ws = []
    for sheet in wb:
        ws.append(sheet)

    #escrevendo as variáveis na sheet 2
    for x in range(0,len(dados)):
        ws[2][celulas[x]] = dados[x]

    #salvando o arquivo
    wb.save(f'Solicitação de acesso {dados[0]}.xlsx')

if (__name__ == "__main__"):
    preencher()